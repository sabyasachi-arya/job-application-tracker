"""
Job Application Tracker  –  v4
Fixes:
  • Mouse-wheel scrolling works everywhere (global binding on root)
  • No layout jank: minimal nesting, stable geometry, no pack/grid mix
  • Pages shown/hidden with place() – no flicker, no resize artifacts
  • Config stored next to .exe, never inside PyInstaller bundle
  • Auto-connects if Excel sits next to .exe on first launch
"""

import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from openpyxl import load_workbook
from datetime import datetime, date
import os, sys, json, calendar as _cal

# ── App-directory helper (works as .py AND as .exe) ──────────────────────────
def _app_dir():
    return os.path.dirname(sys.executable) if getattr(sys, "frozen", False) \
           else os.path.dirname(os.path.abspath(__file__))

_CFG = os.path.join(_app_dir(), "config.json")

def _load_cfg():
    try:
        if os.path.exists(_CFG):
            with open(_CFG) as f: return json.load(f)
    except Exception: pass
    return {}

def _save_cfg(d):
    with open(_CFG, "w") as f: json.dump(d, f, indent=2)

def get_xl():  return _load_cfg().get("excel_path", "")
def set_xl(p): c = _load_cfg(); c["excel_path"] = p; _save_cfg(c)

# Auto-connect: if no config yet AND Excel sits next to the exe, wire it up
_AUTO = os.path.join(_app_dir(), "Job-application-tracksheet.xlsx")
if not os.path.exists(_CFG) and os.path.exists(_AUTO):
    set_xl(_AUTO)

# ── Excel helpers ─────────────────────────────────────────────────────────────
COLS     = ["Application ID","Date of Application","Company Name","Website",
            "Job Position","Application Done","Call Received","Interview",
            "1st Round","2nd Round","HR Round","Rejected"]
WEBSITES = ["LinkedIn","NaukhriGulf","Indeed","Job Portal","Others"]
APP_DONE = ["Yes","No","Saved"]
YES_NO   = ["No","Yes"]

def _wb(p):     return load_workbook(p)
def _ws(wb):    return wb["Sheet1"]

def all_rows(p):
    ws = _ws(_wb(p))
    return [list(r) for r in ws.iter_rows(min_row=2, values_only=True)
            if any(v is not None for v in r)]

def next_id(p): return len(all_rows(p)) + 1

def add_row(p, data):
    wb = _wb(p); ws = _ws(wb)
    ri = 2
    while ws.cell(ri, 1).value is not None: ri += 1
    for c, v in enumerate(data, 1): ws.cell(ri, c).value = v
    wb.save(p)

def del_row(p, aid):
    wb = _wb(p); ws = _ws(wb)
    for row in ws.iter_rows(min_row=2):
        if row[0].value == aid:
            ws.delete_rows(row[0].row)
            for i, r in enumerate(ws.iter_rows(min_row=2), 1):
                if r[0].value is not None: r[0].value = i
            wb.save(p); return

def upd_row(p, aid, data):
    wb = _wb(p); ws = _ws(wb)
    for row in ws.iter_rows(min_row=2):
        if row[0].value == aid:
            for c, v in enumerate(data, 1): ws.cell(row[0].row, c).value = v
            wb.save(p); return

def rows_by_date(p, d):    return [r for r in all_rows(p) if str(r[1]) == d]
def site_counts(p):
    out = {w: 0 for w in WEBSITES}
    for r in all_rows(p):
        if r[3] in out: out[r[3]] += 1
    return out
def yes_ct(p, col):
    i = COLS.index(col)
    return sum(1 for r in all_rows(p) if str(r[i]) == "Yes")

# ── Colours & fonts ───────────────────────────────────────────────────────────
BG    = "#F0F3FA"
PANEL = "#FFFFFF"
SIDE  = "#1A2847"
SIDH  = "#22355E"
ACC   = "#D94F2B"
ACCH  = "#B83D1F"
BRD   = "#D5DCF0"
TXT   = "#1C2540"
MUT   = "#667299"
INP   = "#F3F5FF"
GRN   = "#1B9E55"
RED   = "#CC2E2E"
AMB   = "#D07600"
VIO   = "#6040E0"
TEA   = "#0A8585"

FT  = ("Segoe UI", 13, "bold")
FH  = ("Segoe UI", 11, "bold")
FB  = ("Segoe UI", 10)
FS  = ("Segoe UI",  9)
FL  = ("Segoe UI",  9, "bold")
FN  = ("Segoe UI", 10, "bold")

# ── Global mouse-wheel scroll ─────────────────────────────────────────────────
# We register ONE binding on the root window and route to whatever canvas
# the cursor is over. This is the only reliable cross-platform approach.

_SCROLL_CANVASES = []   # all live scrollable canvases, in order of creation

def _register_canvas(cv):
    _SCROLL_CANVASES.append(cv)

def _unregister_canvas(cv):
    try: _SCROLL_CANVASES.remove(cv)
    except ValueError: pass

def _global_scroll(event):
    """Find the topmost canvas under the cursor and scroll it."""
    wx, wy = event.x_root, event.y_root
    for cv in reversed(_SCROLL_CANVASES):
        try:
            if not cv.winfo_exists(): continue
            cx, cy = cv.winfo_rootx(), cv.winfo_rooty()
            cw, ch = cv.winfo_width(), cv.winfo_height()
            if cx <= wx <= cx + cw and cy <= wy <= cy + ch:
                delta = 0
                if event.delta:      delta = int(-1 * (event.delta / 120))
                elif event.num == 4: delta = -1
                elif event.num == 5: delta = 1
                if delta: cv.yview_scroll(delta, "units")
                return
        except Exception: pass

# ── Smooth ScrollFrame ────────────────────────────────────────────────────────
class SF(tk.Frame):
    """
    Vertically scrollable frame.
    Mouse-wheel handled globally via _global_scroll — no per-widget binding needed.
    """
    def __init__(self, parent, bg=BG, **kw):
        super().__init__(parent, bg=bg, **kw)
        self.cv  = tk.Canvas(self, bg=bg, highlightthickness=0, bd=0)
        self.vsb = ttk.Scrollbar(self, orient="vertical", command=self.cv.yview)
        self.cv.configure(yscrollcommand=self.vsb.set)
        self.vsb.pack(side="right", fill="y")
        self.cv.pack(side="left",  fill="both", expand=True)

        self.body = tk.Frame(self.cv, bg=bg)
        self._win = self.cv.create_window((0, 0), window=self.body, anchor="nw")

        self.body.bind("<Configure>",
                       lambda e: self.cv.configure(scrollregion=self.cv.bbox("all")))
        self.cv.bind("<Configure>",
                     lambda e: self.cv.itemconfig(self._win, width=e.width))

        _register_canvas(self.cv)
        self.bind("<Destroy>", lambda e: _unregister_canvas(self.cv))

    def scroll_top(self): self.cv.yview_moveto(0)


# ── Scrollable Treeview frame ─────────────────────────────────────────────────
class TreeFrame(tk.Frame):
    """Treeview + scrollbars; mouse-wheel handled globally."""
    def __init__(self, parent, columns, col_widths, **kw):
        super().__init__(parent, bg=BG, **kw)
        self.tree = ttk.Treeview(self, columns=columns, show="headings")
        vsb = ttk.Scrollbar(self, orient="vertical",   command=self.tree.yview)
        hsb = ttk.Scrollbar(self, orient="horizontal", command=self.tree.xview)
        self.tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)

        for col, w in zip(columns, col_widths):
            self.tree.heading(col, text=col)
            self.tree.column(col, width=w, minwidth=46, anchor="w")

        self.tree.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        hsb.grid(row=1, column=0, sticky="ew")
        self.rowconfigure(0, weight=1); self.columnconfigure(0, weight=1)
        self.tree.tag_configure("odd",  background="#F6F8FF")
        self.tree.tag_configure("even", background=PANEL)

        # Register as scrollable target
        _register_canvas(self.tree)
        self.bind("<Destroy>", lambda e: _unregister_canvas(self.tree))


# ── Widget factories ──────────────────────────────────────────────────────────
def _entry(parent, tv=None, w=22):
    e = tk.Entry(parent, textvariable=tv, width=w, font=FB,
                 bg=INP, fg=TXT, relief="flat", bd=0,
                 highlightthickness=1, highlightbackground=BRD,
                 highlightcolor=ACC, insertbackground=TXT)
    e.bind("<FocusIn>",  lambda ev: e.config(highlightbackground=ACC))
    e.bind("<FocusOut>", lambda ev: e.config(highlightbackground=BRD))
    return e

def _combo(parent, values, w=16):
    cb = ttk.Combobox(parent, values=values, state="readonly",
                      width=w, font=FB, style="App.TCombobox")
    return cb

def _btn(parent, text, cmd, bg=ACC, fg="white", px=18, py=8):
    def darker(h):
        try: r,g,b = int(h[1:3],16),int(h[3:5],16),int(h[5:7],16)
        except Exception: return h
        return f"#{max(r-28,0):02x}{max(g-28,0):02x}{max(b-28,0):02x}"
    dark = darker(bg)
    b = tk.Button(parent, text=text, command=cmd, bg=bg, fg=fg,
                  font=FL, relief="flat", bd=0, cursor="hand2",
                  activebackground=dark, activeforeground=fg,
                  padx=px, pady=py)
    b.bind("<Enter>", lambda e: b.config(bg=dark))
    b.bind("<Leave>", lambda e: b.config(bg=bg))
    return b

def _hr(parent, pady=10):
    tk.Frame(parent, bg=BRD, height=1).pack(fill="x", pady=pady)

def _cap(parent, text):
    tk.Label(parent, text=text.upper(), bg=parent["bg"], fg=MUT,
             font=("Segoe UI", 8, "bold")).pack(anchor="w", pady=(0, 4))


# ── Calendar popup ─────────────────────────────────────────────────────────────
class CalPop(tk.Toplevel):
    def __init__(self, anchor, callback):
        top = anchor.winfo_toplevel()
        super().__init__(top)
        # NO overrideredirect — it breaks grab_set on Windows and causes crashes
        self.title("Pick a Date")
        self.resizable(False, False)
        self.configure(bg=BRD)
        self.cb = callback
        t = date.today()
        self.yr, self.mo, self._sel = t.year, t.month, t
        self._draw()
        # Position near the anchor widget, then grab focus
        self.update_idletasks()
        ax = anchor.winfo_rootx()
        ay = anchor.winfo_rooty() + anchor.winfo_height() + 6
        sw = self.winfo_screenwidth(); sh = self.winfo_screenheight()
        pw = self.winfo_reqwidth(); ph = self.winfo_reqheight()
        self.geometry(f"+{min(ax, sw-pw-10)}+{min(ay, sh-ph-10)}")
        self.transient(top)
        self.grab_set()     # grab AFTER geometry is set — prevents Windows crash
        self.focus_force()

    def _draw(self):
        for w in self.winfo_children(): w.destroy()
        wrap = tk.Frame(self, bg=PANEL); wrap.pack(fill="both", expand=True)

        hdr = tk.Frame(wrap, bg=SIDE, pady=9); hdr.pack(fill="x")
        tk.Button(hdr, text="‹", command=self._prev, bg=SIDE, fg="white",
                  font=("Segoe UI",13), relief="flat", cursor="hand2",
                  activebackground=SIDH).pack(side="left", padx=8)
        tk.Label(hdr, text=f"{_cal.month_name[self.mo]}  {self.yr}",
                 bg=SIDE, fg="white", font=FH).pack(side="left", expand=True)
        tk.Button(hdr, text="›", command=self._next, bg=SIDE, fg="white",
                  font=("Segoe UI",13), relief="flat", cursor="hand2",
                  activebackground=SIDH).pack(side="right", padx=8)

        grid = tk.Frame(wrap, bg=PANEL, padx=10, pady=8); grid.pack()
        for i, d in enumerate(["Mo","Tu","We","Th","Fr","Sa","Su"]):
            tk.Label(grid, text=d, width=3, bg=BG, fg=MUT,
                     font=("Segoe UI",8,"bold")).grid(row=0, column=i, padx=1, pady=1)

        today = date.today()
        for r, week in enumerate(_cal.monthcalendar(self.yr, self.mo), 1):
            for c, day in enumerate(week):
                if day == 0:
                    tk.Label(grid, text="", width=3,
                             bg=PANEL).grid(row=r, column=c, padx=1, pady=1)
                    continue
                is_sel   = (day==self._sel.day and self.mo==self._sel.month
                            and self.yr==self._sel.year)
                is_today = (day==today.day and self.mo==today.month
                            and self.yr==today.year)
                bg_ = ACC if is_sel else (INP if is_today else PANEL)
                fg_ = "white" if is_sel else (ACC if is_today else TXT)
                tk.Button(grid, text=str(day), width=3, bg=bg_, fg=fg_,
                          relief="flat", cursor="hand2", font=("Segoe UI",9),
                          activebackground=ACC, activeforeground="white",
                          command=lambda d=day: self._pick(d)
                          ).grid(row=r, column=c, padx=1, pady=1)

        tk.Button(wrap, text="✕  Close", command=self.destroy,
                  bg=BG, fg=MUT, font=FS, relief="flat",
                  cursor="hand2", pady=5).pack(fill="x")

    def _prev(self):
        self.mo -= 1
        if self.mo == 0: self.mo, self.yr = 12, self.yr-1
        self._draw()

    def _next(self):
        self.mo += 1
        if self.mo == 13: self.mo, self.yr = 1, self.yr+1
        self._draw()

    def _pick(self, day):
        self._sel = date(self.yr, self.mo, day)
        self.cb(self._sel.strftime("%d-%m-%Y"))
        self.destroy()


# ── Main window ───────────────────────────────────────────────────────────────
class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Job Application Tracker")
        self.geometry("1240x760"); self.minsize(980, 640)
        self.configure(bg=BG)

        # Global mouse-wheel — one binding on the root catches everything
        self.bind_all("<MouseWheel>", _global_scroll)
        self.bind_all("<Button-4>",   _global_scroll)
        self.bind_all("<Button-5>",   _global_scroll)

        self._xl   = tk.StringVar(value=get_xl())
        self._cur  = None
        self._pgs  = {}

        self._styles()
        self._layout()

        if self._xl.get() and os.path.exists(self._xl.get()):
            self._reload_all()
        self._nav("add")

    # ── Styles ────────────────────────────────────────────────────────────────
    def _styles(self):
        s = ttk.Style(); s.theme_use("clam")
        s.configure("App.TCombobox",
                    fieldbackground=INP, background=INP, foreground=TXT,
                    selectbackground=ACC, relief="flat", borderwidth=1,
                    arrowsize=14, padding=4)
        s.map("App.TCombobox",
              fieldbackground=[("readonly", INP)],
              selectbackground=[("readonly", ACC)])
        s.configure("Treeview",
                    background=PANEL, foreground=TXT,
                    rowheight=32, fieldbackground=PANEL, font=FB, borderwidth=0)
        s.configure("Treeview.Heading",
                    background=SIDE, foreground="white",
                    font=FL, relief="flat", padding=8)
        s.map("Treeview",
              background=[("selected", ACC)],
              foreground=[("selected", "white")])
        for ori in ("Vertical", "Horizontal"):
            s.configure(f"{ori}.TScrollbar",
                        background=BRD, troughcolor=BG,
                        borderwidth=0, arrowsize=11, relief="flat")

    # ── Layout ────────────────────────────────────────────────────────────────
    def _layout(self):
        # Sidebar — fixed width, never shrinks
        self._sb = tk.Frame(self, bg=SIDE, width=234)
        self._sb.pack(side="left", fill="y"); self._sb.pack_propagate(False)

        # Main area
        self._main = tk.Frame(self, bg=BG)
        self._main.pack(side="left", fill="both", expand=True)

        self._sidebar()

        # Top bar
        tbar = tk.Frame(self._main, bg=PANEL, height=54)
        tbar.pack(fill="x"); tbar.pack_propagate(False)
        tk.Frame(tbar, bg=BRD, height=1).pack(side="bottom", fill="x")
        self._ttl = tk.Label(tbar, text="", bg=PANEL, fg=TXT, font=FT)
        self._ttl.pack(side="left", padx=26, pady=12)
        tk.Label(tbar,
                 text=f"📅  {date.today().strftime('%A, %d %B %Y')}",
                 bg=BG, fg=MUT, font=FS, padx=12, pady=5
                 ).pack(side="right", padx=18)

        # Page area — pages use place() so there's zero reflow on switch
        self._area = tk.Frame(self._main, bg=BG)
        self._area.pack(fill="both", expand=True)
        self._area.bind("<Configure>", self._resize_pages)

        self._page_add()
        self._page_search()
        self._page_view()
        self._page_analytics()

    def _resize_pages(self, e):
        for pg in self._pgs.values():
            pg.place_configure(width=e.width, height=e.height)

    # ── Sidebar ───────────────────────────────────────────────────────────────
    def _sidebar(self):
        logo = tk.Frame(self._sb, bg=SIDH, pady=22); logo.pack(fill="x")
        tk.Label(logo, text="💼", font=("Segoe UI",24), bg=SIDH, fg="white").pack()
        tk.Label(logo, text="Job Tracker",
                 font=("Segoe UI",15,"bold"), bg=SIDH, fg="white").pack()
        tk.Label(logo, text="Application Manager",
                 font=("Segoe UI",8), bg=SIDH, fg="#7A9CC6").pack(pady=(2,0))
        tk.Frame(self._sb, bg="#283D6B", height=1).pack(fill="x")

        self._nb = {}
        nav_wrap = tk.Frame(self._sb, bg=SIDE); nav_wrap.pack(fill="x", pady=10)
        for key, icon, lbl in [
            ("add",       "➕", "Add Application"),
            ("search",    "🔍", "Search by Date"),
            ("view",      "📊", "All Applications"),
            ("analytics", "📈", "Analytics"),
        ]:
            row   = tk.Frame(nav_wrap, bg=SIDE, cursor="hand2")
            row.pack(fill="x", padx=8, pady=2)
            inn   = tk.Frame(row, bg=SIDE, padx=14, pady=12); inn.pack(fill="x")
            ik    = tk.Label(inn, text=icon, bg=SIDE, fg="white",
                             font=("Segoe UI",12)); ik.pack(side="left")
            lb    = tk.Label(inn, text=lbl, bg=SIDE, fg="#A5BEDD", font=FN)
            lb.pack(side="left", padx=10)
            ws    = [row, inn, ik, lb]
            self._nb[key] = ws

            def _click(k=key): self._nav(k)
            def _ent(e, k=key, ww=ws):
                if self._cur != k:
                    for w in ww: w.configure(bg=SIDH)
            def _lve(e, k=key, ww=ws):
                if self._cur != k:
                    for w in ww: w.configure(bg=SIDE)
            for w in ws:
                w.bind("<Button-1>", lambda e, k=key: _click(k))
                w.bind("<Enter>", _ent); w.bind("<Leave>", _lve)

        # ── File section at bottom ────────────────────────────────────────────
        tk.Frame(self._sb, bg="#283D6B", height=1).pack(side="bottom", fill="x")
        fsec = tk.Frame(self._sb, bg=SIDE, pady=10)
        fsec.pack(side="bottom", fill="x")
        tk.Label(fsec, text="  CONNECTED FILE",
                 font=("Segoe UI",7,"bold"), bg=SIDE, fg="#3D5580").pack(anchor="w")
        self._flbl = tk.Label(fsec, text=self._fn(self._xl.get()),
                               bg=SIDH, fg="#88AACC",
                               font=("Segoe UI",8), anchor="w",
                               padx=8, pady=6, wraplength=208, justify="left")
        self._flbl.pack(fill="x", padx=8, pady=(4,6))
        bb = tk.Frame(fsec, bg=SIDE); bb.pack(fill="x", padx=8)
        tk.Button(bb, text="📂  Browse", command=self._browse,
                  bg=SIDH, fg="#88BBFF", font=("Segoe UI",8),
                  relief="flat", cursor="hand2", padx=8, pady=5,
                  activebackground=SIDH).pack(side="left")
        tk.Button(bb, text="Connect ✓", command=self._connect,
                  bg=GRN, fg="white", font=("Segoe UI",8,"bold"),
                  relief="flat", cursor="hand2", padx=10, pady=5,
                  activebackground="#157A40").pack(side="left", padx=(6,0))

    def _fn(self, p):
        return f"📄  {os.path.basename(p)}" if p else "⚠  No file connected"

    # ── Navigation ────────────────────────────────────────────────────────────
    _TITLES = {"add":"Add Application","search":"Search by Date",
               "view":"All Applications","analytics":"Analytics"}

    def _nav(self, key):
        if self._cur == key: return
        self._cur = key
        self._ttl.configure(text=self._TITLES.get(key, ""))

        for k, ws in self._nb.items():
            act = (k == key)
            for w in ws: w.configure(bg=ACC if act else SIDE)
            ws[3].configure(fg="white" if act else "#A5BEDD")   # label

        for pg in self._pgs.values(): pg.place_forget()
        pg = self._pgs[key]
        pg.place(in_=self._area, x=0, y=0,
                 width=self._area.winfo_width() or 960,
                 height=self._area.winfo_height() or 700)

        # Lazy data loads
        if key == "add":       self._refresh_id()
        if key == "view":      self._load_view()
        if key == "analytics": self._load_analytics()

    def _browse(self):
        p = filedialog.askopenfilename(
            title="Select your Excel file",
            filetypes=[("Excel Workbook","*.xlsx"), ("All files","*.*")])
        if p: self._xl.set(p); self._flbl.configure(text=self._fn(p))

    def _connect(self):
        p = self._xl.get()
        if not p or not os.path.exists(p):
            messagebox.showerror("File not found",
                "Use Browse to select your Excel file first."); return
        set_xl(p)
        self._flbl.configure(text=self._fn(p))
        self._reload_all()
        messagebox.showinfo("Connected ✓",
            f"Connected to:\n{os.path.basename(p)}\n\n"
            "Every change saves directly to this file.\n"
            "No import or export ever needed.")

    def _reload_all(self):
        self._refresh_id(); self._load_view(); self._load_analytics()

    def _ok(self):
        p = self._xl.get()
        if p and os.path.exists(p): return p
        messagebox.showerror("No file connected",
            "Use Browse → Connect in the sidebar first.")
        return None

    # ─────────────────────────────────────────────────────────────────────────
    # PAGE: ADD APPLICATION
    # ─────────────────────────────────────────────────────────────────────────
    def _page_add(self):
        pg = tk.Frame(self._area, bg=BG); self._pgs["add"] = pg

        sf = SF(pg, bg=BG); sf.pack(fill="both", expand=True)
        b  = sf.body   # everything goes into sf.body

        # outer padding
        pad = tk.Frame(b, bg=BG); pad.pack(fill="x", padx=34, pady=24)

        card = tk.Frame(pad, bg=PANEL, highlightthickness=1,
                        highlightbackground=BRD); card.pack(fill="x")

        # Card header
        ch = tk.Frame(card, bg=ACC, pady=16); ch.pack(fill="x")
        tk.Label(ch, text="  ➕   New Job Application",
                 bg=ACC, fg="white", font=FT).pack(side="left", padx=20)

        body = tk.Frame(card, bg=PANEL, padx=32, pady=22); body.pack(fill="x")

        # ── ID ────────────────────────────────────────────────────────────────
        idf = tk.Frame(body, bg=PANEL); idf.pack(fill="x")
        _cap(idf, "Application ID  (auto-assigned)")
        self._add_id = tk.StringVar(value="—")
        tk.Label(idf, textvariable=self._add_id, bg=ACC, fg="white",
                 font=("Segoe UI",20,"bold"), padx=18, pady=6,
                 width=5, anchor="center").pack(anchor="w", pady=(0,4))

        _hr(body, 14)

        # ── Date ──────────────────────────────────────────────────────────────
        _cap(body, "Date of Application")
        dr = tk.Frame(body, bg=PANEL); dr.pack(anchor="w", pady=(0, 4))
        self._add_date = tk.StringVar(value=date.today().strftime("%d-%m-%Y"))
        de = _entry(dr, tv=self._add_date, w=14); de.pack(side="left", ipady=6)
        _btn(dr, "📅  Pick Date", lambda: CalPop(de, self._add_date.set),
             bg=INP, fg=ACC, px=12, py=6).pack(side="left", padx=10)
        tk.Label(dr, text="DD-MM-YYYY", bg=PANEL, fg=MUT, font=FS).pack(side="left")

        _hr(body, 14)

        # ── 2-col rows ────────────────────────────────────────────────────────
        def two(l1, m1, l2, m2):
            fr = tk.Frame(body, bg=PANEL); fr.pack(fill="x", pady=6)
            for label, maker, side, px in [(l1,m1,"left",(0,18)),
                                            (l2,m2,"right",0)]:
                col = tk.Frame(fr, bg=PANEL)
                col.pack(side=side, expand=True, fill="x", padx=px)
                _cap(col, label)
                maker(col).pack(fill="x", ipady=5)

        def ref(attr, fn):
            def mk(p): w = fn(p); setattr(self, attr, w); return w
            return mk

        def cbo(attr, vals, w=18, idx=0):
            def mk(p):
                cb = _combo(p, vals, w); cb.current(idx)
                setattr(self, attr, cb); return cb
            return mk

        two("Company Name",    ref("_co", lambda p: _entry(p, w=30)),
            "Website / Portal", cbo("_ws", WEBSITES))
        two("Job Position",    ref("_po", lambda p: _entry(p, w=30)),
            "Application Done", cbo("_ad", APP_DONE))

        _hr(body, 14)

        # ── Status grid ───────────────────────────────────────────────────────
        _cap(body, "Recruitment Status  —  all default to No")
        sg = tk.Frame(body, bg=PANEL); sg.pack(fill="x")
        sg.columnconfigure((0,1,2), weight=1)

        for i, (icon, label, attr) in enumerate([
            ("📞","Call Received","_sc"), ("🗣","Interview","_si"),
            ("1️⃣","1st Round","_s1"),    ("2️⃣","2nd Round","_s2"),
            ("🤝","HR Round","_sh"),      ("❌","Rejected","_sr"),
        ]):
            cell = tk.Frame(sg, bg=BG, highlightthickness=1,
                            highlightbackground=BRD, padx=12, pady=12)
            cell.grid(row=i//3, column=i%3, padx=5, pady=5, sticky="ew")
            tk.Label(cell, text=f"{icon}  {label}",
                     bg=BG, fg=TXT, font=FB).pack(anchor="w")
            cb = _combo(cell, YES_NO, 8); cb.current(0)
            cb.pack(anchor="w", pady=(5,0), ipady=3)
            setattr(self, attr, cb)

        _hr(body, 14)

        # ── Submit row ────────────────────────────────────────────────────────
        sr = tk.Frame(body, bg=PANEL); sr.pack(fill="x")
        _btn(sr, "  ✅   Submit & Save to Excel  ",
             self._submit, bg=GRN, px=24, py=10).pack(side="left")
        _btn(sr, "  🔄   Reset Form  ",
             self._reset_add, bg=MUT, px=16, py=10).pack(side="left", padx=12)

    def _refresh_id(self):
        p = self._xl.get()
        self._add_id.set(str(next_id(p)) if p and os.path.exists(p) else "—")

    def _reset_add(self):
        self._co.delete(0, "end"); self._po.delete(0, "end")
        self._add_date.set(date.today().strftime("%d-%m-%Y"))
        self._ws.current(0); self._ad.current(0)
        for a in ("_sc","_si","_s1","_s2","_sh","_sr"):
            getattr(self, a).current(0)
        self._refresh_id()

    def _submit(self):
        p = self._ok()
        if not p: return
        co = self._co.get().strip(); po = self._po.get().strip()
        dv = self._add_date.get().strip()
        if not co: messagebox.showwarning("Missing field","Enter a Company Name."); return
        if not po: messagebox.showwarning("Missing field","Enter a Job Position."); return
        try: datetime.strptime(dv, "%d-%m-%Y")
        except ValueError:
            messagebox.showerror("Invalid date","Date must be DD-MM-YYYY."); return

        nid = next_id(p)
        add_row(p, [nid, dv, co, self._ws.get(), po, self._ad.get(),
                    self._sc.get(), self._si.get(), self._s1.get(),
                    self._s2.get(), self._sh.get(), self._sr.get()])

        pop = tk.Toplevel(self); pop.title("Saved")
        pop.resizable(False,False); pop.configure(bg=PANEL)
        pop.grab_set(); pop.transient(self); pop.geometry("320x190")
        tk.Label(pop, text="✅", font=("Segoe UI",36), bg=PANEL).pack(pady=(18,4))
        tk.Label(pop, text="Application Saved!",
                 bg=PANEL, fg=GRN, font=FT).pack()
        tk.Label(pop, text=f"#{nid}  ·  {co}  ·  {po}",
                 bg=PANEL, fg=MUT, font=FB).pack(pady=(4,14))
        _btn(pop,"  Add Another  ",
             lambda: (pop.destroy(), self._reset_add()), px=18, py=7).pack()
        self.after(2600, lambda: pop.destroy() if pop.winfo_exists() else None)
        self._reset_add(); self._load_view(); self._load_analytics()

    # ─────────────────────────────────────────────────────────────────────────
    # PAGE: SEARCH BY DATE
    # ─────────────────────────────────────────────────────────────────────────
    def _page_search(self):
        pg = tk.Frame(self._area, bg=BG); self._pgs["search"] = pg

        # Fixed top search bar
        top = tk.Frame(pg, bg=PANEL, highlightthickness=1,
                       highlightbackground=BRD)
        top.pack(fill="x", padx=28, pady=(20, 0))

        left = tk.Frame(top, bg=PANEL, padx=20, pady=16)
        left.pack(side="left")
        tk.Label(left, text="Find applications for a date:",
                 bg=PANEL, fg=MUT, font=FS).pack(anchor="w", pady=(0,5))
        dr = tk.Frame(left, bg=PANEL); dr.pack(anchor="w")
        self._sd = tk.StringVar(value=date.today().strftime("%d-%m-%Y"))
        se = _entry(dr, tv=self._sd, w=14); se.pack(side="left", ipady=6)
        _btn(dr, "📅", lambda: CalPop(se, self._sd.set),
             bg=INP, fg=ACC, px=8, py=6).pack(side="left", padx=6)
        tk.Label(dr, text="DD-MM-YYYY", bg=PANEL, fg=MUT, font=FS).pack(side="left", padx=6)

        _btn(top, "  🔍  Search  ", self._do_search,
             px=22, py=10).pack(side="right", padx=20)

        # Results area — scrollable
        self._sr_sf = SF(pg, bg=BG)
        self._sr_sf.pack(fill="both", expand=True, padx=28, pady=10)

    def _do_search(self):
        p = self._ok()
        if not p: return
        d = self._sd.get().strip()
        try: datetime.strptime(d, "%d-%m-%Y")
        except ValueError:
            messagebox.showerror("Invalid date","Use DD-MM-YYYY format."); return

        # Clear previous results
        for w in self._sr_sf.body.winfo_children(): w.destroy()
        self._sr_sf.scroll_top()

        results = rows_by_date(p, d)
        body    = self._sr_sf.body   # parent for all result cards

        if not results:
            empty = tk.Frame(body, bg=PANEL,
                             highlightthickness=1, highlightbackground=BRD)
            empty.pack(fill="x", pady=8)
            tk.Label(empty, text="🔍", font=("Segoe UI",30), bg=PANEL).pack(pady=(18,6))
            tk.Label(empty, text=f"No applications found for  {d}",
                     bg=PANEL, fg=MUT, font=FH).pack(pady=(0,18))
        else:
            banner = tk.Frame(body, bg=ACC)
            banner.pack(fill="x", pady=(0,6))
            tk.Label(banner,
                     text=f"  {len(results)}  result(s)  for  {d}  — "
                          "click Edit or Remove on any row",
                     bg=ACC, fg="white", font=FH).pack(side="left", padx=14, pady=9)
            for row in results:
                self._result_card(body, row, p)

    def _result_card(self, parent, row, path):
        card = tk.Frame(parent, bg=PANEL,
                        highlightthickness=1, highlightbackground=BRD)
        card.pack(fill="x", pady=4, padx=1)

        # Accent left bar
        tk.Frame(card, bg=ACC, width=5).pack(side="left", fill="y")

        body = tk.Frame(card, bg=PANEL, pady=11, padx=16)
        body.pack(side="left", fill="x", expand=True)

        # Row 1: id badge + company + position
        r1 = tk.Frame(body, bg=PANEL); r1.pack(fill="x")
        tk.Label(r1, text=f" #{row[0]} ", bg=ACC, fg="white",
                 font=("Segoe UI",10,"bold"), padx=4, pady=2).pack(side="left")
        tk.Label(r1, text=f"  {row[2]}", bg=PANEL, fg=TXT,
                 font=FH).pack(side="left")
        tk.Label(r1, text=f"  ·  {row[4]}", bg=PANEL, fg=MUT,
                 font=FB).pack(side="left")

        # Row 2: status pills
        r2 = tk.Frame(body, bg=PANEL); r2.pack(fill="x", pady=(5,0))

        def pill(txt, hi=False, bad=False):
            bg_ = RED if bad else (GRN if hi else BG)
            fg_ = "white" if (hi or bad) else MUT
            tk.Label(r2, text=txt, bg=bg_, fg=fg_, font=FS,
                     padx=7, pady=3).pack(side="left", padx=(0,5))

        pill(f"📅 {row[1]}")
        pill(f"🌐 {row[3]}")
        pill(f"App: {row[5]}",  hi=(row[5]=="Yes"))
        pill(f"Call: {row[6]}", hi=(row[6]=="Yes"))
        pill(f"Interview: {row[7]}", hi=(row[7]=="Yes"))
        pill(f"R1: {row[8]}",  hi=(row[8]=="Yes"))
        pill(f"R2: {row[9]}",  hi=(row[9]=="Yes"))
        pill(f"HR: {row[10]}", hi=(row[10]=="Yes"))
        pill(f"Rejected: {row[11]}", bad=(row[11]=="Yes"))

        # Buttons
        bc = tk.Frame(card, bg=PANEL, padx=12)
        bc.pack(side="right", fill="y")
        _btn(bc, "✏️  Edit",
             lambda r=row: self._edit(r, path),
             bg=AMB, px=12, py=7).pack(pady=(12,4))
        _btn(bc, "🗑  Remove",
             lambda r=row: self._remove(r, path),
             bg=RED, px=12, py=7).pack(pady=(0,12))

    def _remove(self, row, path):
        w = tk.Toplevel(self); w.title("Confirm Removal")
        w.resizable(False,False); w.configure(bg=PANEL)
        w.grab_set(); w.transient(self); w.geometry("420x234")
        tk.Label(w, text="⚠️", font=("Segoe UI",30), bg=PANEL).pack(pady=(18,4))
        tk.Label(w, text="Remove this application?",
                 bg=PANEL, fg=TXT, font=FT).pack()
        tk.Label(w, text="This cannot be undone.",
                 bg=PANEL, fg=MUT, font=FS).pack(pady=(2,10))
        info = tk.Frame(w, bg=BG, padx=12, pady=8,
                        highlightthickness=1, highlightbackground=BRD)
        info.pack(fill="x", padx=24)
        tk.Label(info,
                 text=f"#{row[0]}  ·  {row[1]}  ·  {row[2]}  ·  {row[4]}",
                 bg=BG, fg=TXT, font=FB).pack()
        bb = tk.Frame(w, bg=PANEL); bb.pack(pady=14)
        _btn(bb, "Cancel", w.destroy, bg=MUT, px=14).pack(side="left", padx=6)
        def go():
            del_row(path, row[0]); w.destroy()
            self._do_search(); self._load_view(); self._load_analytics()
        _btn(bb, "Yes, Remove", go, bg=RED, px=14).pack(side="left")

    def _edit(self, row, path):
        w = tk.Toplevel(self); w.title(f"Edit Application #{row[0]}")
        w.resizable(False,False); w.configure(bg=PANEL)
        w.grab_set(); w.transient(self)

        hdr = tk.Frame(w, bg=SIDE, pady=14); hdr.pack(fill="x")
        tk.Label(hdr, text=f"  ✏️   Edit  Application  #{row[0]}",
                 bg=SIDE, fg="white", font=FT).pack(side="left", padx=18)

        b = tk.Frame(w, bg=PANEL, padx=24, pady=18); b.pack(fill="both")

        def div(): _hr(b, 10)

        # Date
        _cap(b, "Date of Application")
        dv = tk.StringVar(value=str(row[1]))
        dr = tk.Frame(b, bg=PANEL); dr.pack(anchor="w", pady=(0,6))
        de = _entry(dr, tv=dv, w=14); de.pack(side="left", ipady=6)
        _btn(dr, "📅", lambda: CalPop(de, dv.set),
             bg=INP, fg=ACC, px=8, py=6).pack(side="left", padx=6)
        div()

        def two(l1, m1, l2, m2):
            fr = tk.Frame(b, bg=PANEL); fr.pack(fill="x", pady=6)
            for label, maker, side, px in [(l1,m1,"left",(0,18)),
                                            (l2,m2,"right",0)]:
                col = tk.Frame(fr, bg=PANEL)
                col.pack(side=side, expand=True, fill="x", padx=px)
                _cap(col, label)
                maker(col).pack(fill="x", ipady=5)

        cv = tk.StringVar(value=str(row[2]))
        pv = tk.StringVar(value=str(row[4]))
        wr=[]; ar=[]

        def mk_co(p): return _entry(p, tv=cv, w=24)
        def mk_we(p):
            cb=_combo(p,WEBSITES,18)
            cb.current(WEBSITES.index(row[3]) if row[3] in WEBSITES else 0)
            wr.append(cb); return cb
        def mk_po(p): return _entry(p, tv=pv, w=24)
        def mk_ap(p):
            cb=_combo(p,APP_DONE,14)
            cb.current(APP_DONE.index(str(row[5])) if str(row[5]) in APP_DONE else 0)
            ar.append(cb); return cb

        two("Company Name",mk_co,"Website",mk_we)
        two("Job Position",mk_po,"Application Done",mk_ap)
        div()

        _cap(b, "Recruitment Status")
        yg = tk.Frame(b, bg=PANEL); yg.pack(fill="x")
        yg.columnconfigure((0,1,2), weight=1)
        yn = {}
        for i,(lbl,val) in enumerate([
            ("Call Received",row[6]),("Interview",row[7]),
            ("1st Round",row[8]),("2nd Round",row[9]),
            ("HR Round",row[10]),("Rejected",row[11])
        ]):
            cell = tk.Frame(yg, bg=BG, highlightthickness=1,
                            highlightbackground=BRD, padx=12, pady=10)
            cell.grid(row=i//3, column=i%3, padx=4, pady=4, sticky="ew")
            tk.Label(cell, text=lbl, bg=BG, fg=TXT, font=FB).pack(anchor="w")
            cb = _combo(cell, YES_NO, 8)
            cb.current(YES_NO.index(str(val)) if str(val) in YES_NO else 0)
            cb.pack(anchor="w", pady=(4,0), ipady=3)
            yn[lbl] = cb
        div()

        def confirm():
            dval = dv.get().strip()
            try: datetime.strptime(dval, "%d-%m-%Y")
            except ValueError:
                messagebox.showerror("Invalid date","Use DD-MM-YYYY",parent=w); return
            if not messagebox.askyesno("Confirm","Save these changes?",parent=w): return
            upd_row(path, row[0],
                [row[0], dval, cv.get().strip(),
                 wr[0].get() if wr else row[3],
                 pv.get().strip(),
                 ar[0].get() if ar else row[5],
                 yn["Call Received"].get(), yn["Interview"].get(),
                 yn["1st Round"].get(), yn["2nd Round"].get(),
                 yn["HR Round"].get(), yn["Rejected"].get()])
            w.destroy()
            self._do_search(); self._load_view(); self._load_analytics()

        _btn(b, "  ✅   Confirm & Save Changes  ",
             confirm, bg=GRN, px=22, py=10).pack(pady=(0,4))

    # ─────────────────────────────────────────────────────────────────────────
    # PAGE: ALL APPLICATIONS
    # ─────────────────────────────────────────────────────────────────────────
    def _page_view(self):
        pg = tk.Frame(self._area, bg=BG); self._pgs["view"] = pg

        # Top bar
        top = tk.Frame(pg, bg=PANEL, highlightthickness=1,
                       highlightbackground=BRD)
        top.pack(fill="x", padx=28, pady=(20,0))
        tk.Label(top, text="All Applications",
                 bg=PANEL, fg=TXT, font=FT).pack(side="left", padx=20, pady=12)
        self._vcnt = tk.Label(top, text="", bg=ACC, fg="white",
                               font=FL, padx=10, pady=3)
        self._vcnt.pack(side="left", padx=6)
        _btn(top, "🔄  Refresh", self._load_view,
             bg=BG, fg=ACC, px=12, py=6).pack(side="right", padx=20)

        # Hint bar
        hint = tk.Frame(pg, bg="#FFFBE6", highlightthickness=1,
                        highlightbackground="#F0D060")
        hint.pack(fill="x", padx=28, pady=(8,0))
        tk.Label(hint, text="💡  Click any row to Edit or Remove it",
                 bg="#FFFBE6", fg="#7A5800", font=FS,
                 padx=14, pady=6).pack(side="left")

        tf = TreeFrame(pg, COLS,
                       [70,110,158,103,158,108,88,78,68,68,68,68])
        tf.pack(fill="both", expand=True, padx=28, pady=(6,14))
        self._tree = tf.tree

        # Floating context menu (hidden until a row is clicked)
        self._ctx = tk.Frame(self, bg=PANEL,
                             highlightthickness=1, highlightbackground=BRD,
                             relief="flat")
        self._ctx_visible = False

        def _show_ctx(event):
            """Show Edit/Remove menu under the clicked row."""
            row_id = self._tree.identify_row(event.y)
            if not row_id:
                _hide_ctx(); return

            # Highlight selected row
            self._tree.selection_set(row_id)
            values = self._tree.item(row_id, "values")
            if not values: return

            # Convert treeview strings back to a list matching COLS
            row_data = list(values)
            # app_id is stored as string in tree, convert to int for del/upd
            try:    row_data[0] = int(row_data[0])
            except: pass

            # Build / rebuild the menu content
            for w in self._ctx.winfo_children(): w.destroy()

            # Title strip
            hd = tk.Frame(self._ctx, bg=SIDE, pady=6); hd.pack(fill="x")
            tk.Label(hd,
                     text=f"  #{row_data[0]}  ·  {row_data[2]}  ·  {row_data[4]}",
                     bg=SIDE, fg="white", font=FL,
                     padx=8).pack(side="left")

            # Edit button
            def do_edit():
                _hide_ctx()
                p = self._ok()
                if p: self._edit(row_data, p)

            edit_btn = tk.Frame(self._ctx, bg=PANEL, cursor="hand2")
            edit_btn.pack(fill="x")
            edit_inner = tk.Frame(edit_btn, bg=PANEL, padx=16, pady=10)
            edit_inner.pack(fill="x")
            tk.Label(edit_inner, text="✏️", bg=PANEL,
                     font=("Segoe UI",13)).pack(side="left")
            tk.Label(edit_inner, text="  Edit this row",
                     bg=PANEL, fg=TXT, font=FB).pack(side="left")

            def _edit_enter(e):
                edit_btn.configure(bg="#EEF2FF")
                edit_inner.configure(bg="#EEF2FF")
                for w in edit_inner.winfo_children(): w.configure(bg="#EEF2FF")
            def _edit_leave(e):
                edit_btn.configure(bg=PANEL)
                edit_inner.configure(bg=PANEL)
                for w in edit_inner.winfo_children(): w.configure(bg=PANEL)

            for w in [edit_btn, edit_inner] + list(edit_inner.winfo_children()):
                w.bind("<Button-1>", lambda e: do_edit())
                w.bind("<Enter>", _edit_enter)
                w.bind("<Leave>", _edit_leave)

            # Divider
            tk.Frame(self._ctx, bg=BRD, height=1).pack(fill="x")

            # Remove button
            def do_remove():
                _hide_ctx()
                p = self._ok()
                if p: self._remove(row_data, p)

            rem_btn = tk.Frame(self._ctx, bg=PANEL, cursor="hand2")
            rem_btn.pack(fill="x")
            rem_inner = tk.Frame(rem_btn, bg=PANEL, padx=16, pady=10)
            rem_inner.pack(fill="x")
            tk.Label(rem_inner, text="🗑", bg=PANEL,
                     font=("Segoe UI",13)).pack(side="left")
            tk.Label(rem_inner, text="  Remove this row",
                     bg=PANEL, fg=RED, font=FB).pack(side="left")

            def _rem_enter(e):
                rem_btn.configure(bg="#FFF0F0")
                rem_inner.configure(bg="#FFF0F0")
                for w in rem_inner.winfo_children(): w.configure(bg="#FFF0F0")
            def _rem_leave(e):
                rem_btn.configure(bg=PANEL)
                rem_inner.configure(bg=PANEL)
                for w in rem_inner.winfo_children(): w.configure(bg=PANEL)

            for w in [rem_btn, rem_inner] + list(rem_inner.winfo_children()):
                w.bind("<Button-1>", lambda e: do_remove())
                w.bind("<Enter>", _rem_enter)
                w.bind("<Leave>", _rem_leave)

            # Position menu below the clicked row
            self._ctx.update_idletasks()
            rx = self._tree.winfo_rootx()
            # find y of the row bottom edge
            bbox = self._tree.bbox(row_id)
            if not bbox: return
            ry = self._tree.winfo_rooty() + bbox[1] + bbox[3]

            # Convert root coords to self (main window) coords
            wx = rx - self.winfo_rootx()
            wy = ry - self.winfo_rooty()

            mw = self._ctx.winfo_reqwidth()
            mh = self._ctx.winfo_reqheight()
            sw = self.winfo_width(); sh = self.winfo_height()
            # Keep inside window bounds
            wx = max(4, min(wx, sw - mw - 4))
            wy = max(4, min(wy, sh - mh - 4))

            self._ctx.place(x=wx, y=wy)
            self._ctx.lift()
            self._ctx_visible = True

        def _hide_ctx(event=None):
            if self._ctx_visible:
                self._ctx.place_forget()
                self._ctx_visible = False

        self._tree.bind("<ButtonRelease-1>", _show_ctx)

        # Clicking anywhere else hides the menu
        self._tree.bind("<FocusOut>",  _hide_ctx)
        pg.bind_all("<Escape>",        _hide_ctx)

        # Hide when user scrolls or clicks outside
        self._tree.bind("<MouseWheel>", lambda e: (_hide_ctx(), None))
        self._tree.bind("<Button-4>",   lambda e: (_hide_ctx(), None))
        self._tree.bind("<Button-5>",   lambda e: (_hide_ctx(), None))

    def _load_view(self):
        p = self._xl.get()
        if not p or not os.path.exists(p): return
        # Hide context menu before reload
        if hasattr(self, '_ctx_visible') and self._ctx_visible:
            self._ctx.place_forget(); self._ctx_visible = False
        for i in self._tree.get_children(): self._tree.delete(i)
        rows = all_rows(p)
        for i, row in enumerate(rows):
            self._tree.insert("","end",
                values=[str(v) if v is not None else "" for v in row],
                tags=("even" if i%2==0 else "odd",))
        self._vcnt.configure(text=f"{len(rows)} total")

    # ─────────────────────────────────────────────────────────────────────────
    # PAGE: ANALYTICS
    # ─────────────────────────────────────────────────────────────────────────
    def _page_analytics(self):
        pg = tk.Frame(self._area, bg=BG); self._pgs["analytics"] = pg

        top = tk.Frame(pg, bg=PANEL, highlightthickness=1,
                       highlightbackground=BRD)
        top.pack(fill="x", padx=28, pady=(20,0))
        tk.Label(top, text="Analytics & Statistics",
                 bg=PANEL, fg=TXT, font=FT).pack(side="left", padx=20, pady=12)
        _btn(top,"🔄  Refresh", self._load_analytics,
             bg=BG, fg=ACC, px=12, py=6).pack(side="right", padx=20)

        self._ana = tk.Frame(pg, bg=BG)
        self._ana.pack(fill="both", expand=True, padx=28, pady=16)

    def _load_analytics(self):
        p = self._xl.get()
        if not p or not os.path.exists(p): return
        for w in self._ana.winfo_children(): w.destroy()

        rows  = all_rows(p); total = len(rows)

        # KPI tiles
        tiles = tk.Frame(self._ana, bg=BG); tiles.pack(fill="x", pady=(0,16))
        tiles.columnconfigure((0,1,2,3), weight=1)

        def tile(col, icon, label, val, color):
            t = tk.Frame(tiles, bg=color, padx=20, pady=16)
            t.grid(row=0, column=col,
                   padx=(0,12) if col < 3 else 0, sticky="ew")
            tk.Label(t, text=icon, bg=color, fg="white",
                     font=("Segoe UI",22)).pack(anchor="w")
            tk.Label(t, text=str(val), bg=color, fg="white",
                     font=("Segoe UI",28,"bold")).pack(anchor="w")
            tk.Label(t, text=label, bg=color, fg="white",
                     font=("Segoe UI",9)).pack(anchor="w")

        tile(0,"📋","Total Applications",     total,                  ACC)
        tile(1,"📞","Calls Received",          yes_ct(p,"Call Received"), GRN)
        tile(2,"🗣", "Interviews",              yes_ct(p,"Interview"),    VIO)
        tile(3,"❌","Rejected",                yes_ct(p,"Rejected"),     RED)

        # Charts
        charts = tk.Frame(self._ana, bg=BG); charts.pack(fill="both", expand=True)
        charts.columnconfigure((0,1), weight=1); charts.rowconfigure(0, weight=1)

        def ccard(col, title):
            c = tk.Frame(charts, bg=PANEL,
                         highlightthickness=1, highlightbackground=BRD)
            c.grid(row=0, column=col,
                   padx=(0,12) if col==0 else 0, sticky="nsew")
            tk.Label(c, text=title, bg=PANEL, fg=TXT, font=FH
                     ).pack(anchor="w", padx=18, pady=(14,8))
            tk.Frame(c, bg=BRD, height=1).pack(fill="x")
            return c

        lc = ccard(0, "Applications by Website")
        rc = ccard(1, "Recruitment Pipeline")

        sc = site_counts(p)
        for (site,cnt), color in zip(sc.items(), [ACC,GRN,AMB,VIO,TEA]):
            self._bar(lc, site, cnt, total, color)

        for stage, color in [
            ("Call Received",GRN), ("Interview",ACC),
            ("1st Round",VIO),     ("2nd Round",AMB),
            ("HR Round",TEA),      ("Rejected",RED)
        ]:
            self._bar(rc, stage, yes_ct(p, stage), total, color)

    def _bar(self, parent, label, count, total, color):
        row = tk.Frame(parent, bg=PANEL); row.pack(fill="x", padx=18, pady=6)
        tk.Label(row, text=label, bg=PANEL, fg=TXT,
                 font=FB, width=15, anchor="w").pack(side="left")
        bg_f = tk.Frame(row, bg=BG, height=14, width=180)
        bg_f.pack(side="left", padx=8); bg_f.pack_propagate(False)
        bw = max(2, int((count / max(total,1)) * 180))
        tk.Frame(bg_f, bg=color, height=14, width=bw).place(x=0,y=0,relheight=1)
        pct = f"{count/total*100:.0f}%" if total else "0%"
        tk.Label(row, text=f"{count}  ({pct})", bg=PANEL, fg=MUT,
                 font=FS, width=10).pack(side="left")


# ── Entry point ───────────────────────────────────────────────────────────────
if __name__ == "__main__":
    App().mainloop()
